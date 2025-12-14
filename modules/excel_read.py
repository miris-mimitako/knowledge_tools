"""Excelファイル読み込みとテキスト抽出モジュール"""
import os
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import json
from pydantic import BaseModel, Field


class ExcelRowData(BaseModel):
    """Excel行データのPydanticモデル"""
    row_index: int = Field(..., ge=0, description="行インデックス")
    data: Dict[str, str] = Field(..., description="列名とセル値の辞書")
    text: str = Field(..., description="行のテキスト（全セルを結合）")


class ExcelSheetData(BaseModel):
    """ExcelシートデータのPydanticモデル"""
    row_count: int = Field(..., ge=0, description="行数")
    column_count: int = Field(..., ge=0, description="列数")
    columns: List[str] = Field(..., description="列名のリスト")
    rows: List[ExcelRowData] = Field(..., description="行データのリスト")
    full_text: str = Field(..., description="全テキスト（全セルを結合）")
    word_count: int = Field(..., ge=0, description="単語数")


class ExcelFileData(BaseModel):
    """ExcelファイルデータのPydanticモデル"""
    file_path: str = Field(..., description="ファイルの絶対パス")
    filename: str = Field(..., description="ファイル名")
    sheet_count: int = Field(..., ge=0, description="シート数")
    sheet_names: List[str] = Field(..., description="シート名のリスト")
    sheets: Dict[str, ExcelSheetData] = Field(..., description="シート名とシートデータの辞書")
    extracted_at: str = Field(..., description="抽出日時（ISO形式）")
    
    class Config:
        json_schema_extra = {
            "example": {
                "file_path": "/path/to/file.xlsx",
                "filename": "file.xlsx",
                "sheet_count": 2,
                "sheet_names": ["Sheet1", "Sheet2"],
                "sheets": {},
                "extracted_at": "2024-01-01T00:00:00"
            }
        }


def read_excel_file(file_path: str, sheet_name: Optional[str] = None) -> ExcelFileData:
    """
    Excelファイルを読み込んでテキスト情報を抽出
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: 読み込むシート名（Noneの場合は全シート）
        
    Returns:
        抽出されたテキスト情報のPydanticモデル
        
    Raises:
        FileNotFoundError: ファイルが存在しない場合
        ValueError: ファイルがExcel形式でない場合
    """
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandasが必要です。'pip install pandas openpyxl'でインストールしてください。")
    
    path = Path(file_path)
    
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")
    
    if not path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
        raise ValueError(f"Excelファイルではありません: {file_path}")
    
    try:
        # Excelファイルを読み込む
        if sheet_name:
            # 指定されたシートのみ読み込む
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            sheets_data = {sheet_name: _extract_dataframe_text(df)}
        else:
            # 全シートを読み込む
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            sheets_data = {}
            for sheet in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet)
                sheets_data[sheet] = _extract_dataframe_text(df)
        
        # sheets_dataは既にExcelSheetDataインスタンスの辞書
        return ExcelFileData(
            file_path=str(path.absolute()),
            filename=path.name,
            sheet_count=len(sheets_data),
            sheet_names=list(sheets_data.keys()),
            sheets=sheets_data,
            extracted_at=datetime.now().isoformat()
        )
    except Exception as e:
        raise ValueError(f"Excelファイルの読み込みエラー: {str(e)}")


def _extract_dataframe_text(df) -> ExcelSheetData:
    """
    pandas DataFrameからテキスト情報を抽出
    
    Args:
        df: pandas DataFrame
        
    Returns:
        抽出されたテキスト情報のPydanticモデル
    """
    import pandas as pd
    
    # NaN値を空文字列に変換
    df = df.fillna("")
    
    # 全セルのテキストを抽出
    all_texts = []
    rows_data = []
    
    for index, row in df.iterrows():
        row_texts = []
        row_data = {}
        
        for col in df.columns:
            cell_value = str(row[col]) if pd.notna(row[col]) else ""
            if cell_value.strip():  # 空でない場合のみ追加
                row_texts.append(cell_value)
                row_data[col] = cell_value
        
        if row_texts:  # 空行でない場合のみ追加
            all_texts.extend(row_texts)
            rows_data.append({
                "row_index": int(index),
                "data": row_data,
                "text": " ".join(row_texts)
            })
    
    # 全テキストを結合
    full_text = " ".join(all_texts)
    
    # 行データをPydanticモデルに変換
    rows_models = [
        ExcelRowData(
            row_index=row["row_index"],
            data=row["data"],
            text=row["text"]
        )
        for row in rows_data
    ]
    
    return ExcelSheetData(
        row_count=len(rows_data),
        column_count=len(df.columns),
        columns=list(df.columns),
        rows=rows_models,
        full_text=full_text,
        word_count=len(full_text.split()) if full_text else 0
    )


def read_excel_simple(file_path: str, sheet_name: Optional[str] = None) -> str:
    """
    Excelファイルを読み込んで、シンプルなテキスト文字列として返す
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: 読み込むシート名（Noneの場合は全シート）
        
    Returns:
        抽出されたテキスト（改行区切り）
    """
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandasが必要です。'pip install pandas openpyxl'でインストールしてください。")
    
    path = Path(file_path)
    
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")
    
    try:
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            return _dataframe_to_text(df)
        else:
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            texts = []
            for sheet in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet)
                sheet_text = f"=== Sheet: {sheet} ===\n{_dataframe_to_text(df)}\n"
                texts.append(sheet_text)
            return "\n".join(texts)
    except Exception as e:
        raise ValueError(f"Excelファイルの読み込みエラー: {str(e)}")


def _dataframe_to_text(df) -> str:
    """
    DataFrameをテキスト形式に変換
    
    Args:
        df: pandas DataFrame
        
    Returns:
        テキスト文字列
    """
    df = df.fillna("")
    lines = []
    
    for index, row in df.iterrows():
        row_values = [str(val) for val in row.values if str(val).strip()]
        if row_values:
            lines.append(" | ".join(row_values))
    
    return "\n".join(lines)


def get_excel_metadata(file_path: str) -> Dict[str, Any]:
    """
    Excelファイルのメタ情報を取得（ファイルを開かずに）
    
    Args:
        file_path: Excelファイルのパス
        
    Returns:
        メタ情報の辞書
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxlが必要です。'pip install openpyxl'でインストールしてください。")
    
    path = Path(file_path)
    
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")
    
    try:
        # ファイルを開いてメタ情報のみ取得
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        sheet_info = []
        total_rows = 0
        total_cols = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            sheet_info.append({
                "name": sheet_name,
                "max_row": max_row,
                "max_column": max_col,
                "active": sheet_name == wb.active.title
            })
            
            total_rows += max_row
            total_cols = max(total_cols, max_col)
        
        wb.close()
        
        return {
            "file_path": str(path.absolute()),
            "filename": path.name,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "active_sheet": wb.active.title if wb.active else None,
            "sheets": sheet_info,
            "total_rows": total_rows,
            "max_columns": total_cols
        }
    except Exception as e:
        raise ValueError(f"Excelファイルのメタ情報取得エラー: {str(e)}")


def extract_excel_to_chunks(file_path: str, chunk_size: int = 1000) -> List[Dict[str, Any]]:
    """
    Excelファイルを読み込んで、チャンク単位でテキストを抽出
    
    Args:
        file_path: Excelファイルのパス
        chunk_size: チャンクサイズ（文字数）
        
    Returns:
        チャンクのリスト
    """
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandasが必要です。'pip install pandas openpyxl'でインストールしてください。")
    
    path = Path(file_path)
    
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {file_path}")
    
    excel_file = pd.ExcelFile(file_path, engine='openpyxl')
    chunks = []
    chunk_id = 0
    
    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        df = df.fillna("")
        
        current_chunk = ""
        current_chunk_rows = []
        
        for index, row in df.iterrows():
            row_text = " ".join([str(val) for val in row.values if str(val).strip()])
            
            if len(current_chunk) + len(row_text) > chunk_size and current_chunk:
                # チャンクを保存
                chunks.append({
                    "chunk_id": chunk_id,
                    "sheet_name": sheet_name,
                    "row_start": current_chunk_rows[0] if current_chunk_rows else index,
                    "row_end": current_chunk_rows[-1] if current_chunk_rows else index,
                    "text": current_chunk.strip(),
                    "char_count": len(current_chunk)
                })
                chunk_id += 1
                current_chunk = ""
                current_chunk_rows = []
            
            current_chunk += row_text + " "
            current_chunk_rows.append(int(index))
        
        # 最後のチャンクを保存
        if current_chunk.strip():
            chunks.append({
                "chunk_id": chunk_id,
                "sheet_name": sheet_name,
                "row_start": current_chunk_rows[0] if current_chunk_rows else 0,
                "row_end": current_chunk_rows[-1] if current_chunk_rows else 0,
                "text": current_chunk.strip(),
                "char_count": len(current_chunk)
            })
            chunk_id += 1
    
    return chunks

